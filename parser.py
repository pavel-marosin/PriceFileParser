from openpyxl import load_workbook
import datetime, sys, os

reload(sys)
sys.setdefaultencoding('utf-8')


class ParseFile():

  _run_date = datetime.datetime.now().date()
  _cwd = os.getcwd()
  _dir_path = os.path.dirname(os.path.realpath(__file__))



  def create_price_file_export(self):

    parsed_array = []
    wb2 = None
    _month_run = ''



    try:
        wb2 = load_workbook(self._cwd + '/source/price_list.xlsx')
    except IOError as e:
        print "I/O error({0}): {1}".format(e.errno, e.strerror) + '. Looks like the price_list.xlxs file may be missing ' \
                                                                  'from the source directory.'
        sys.exit(1)



  #get month of file from user



    sheet_name = wb2.get_sheet_names()

    price_sheet = wb2.get_sheet_by_name(sheet_name[0])
    _num_rows = price_sheet.max_row
    _num_column = price_sheet.max_column


    for x in range(2, _num_rows):
        #initialize an array of 97 blank values
        parsed_list = ['']*97

        if not price_sheet.cell(row=x, column=1).value.startswith('AAASS') \
                and not price_sheet.cell(row=x, column=10).value in ('NS','WH','G','B','R','CL','CS','O','Q','I'):


            discount_index = 37     #start the list of discount codes and values here, relevant to code below

            parsed_list[0] = 'WR'
            parsed_list[1] = str(self._run_date.month + 2).zfill(2)       #current month + 1 to indicate the month file is run for
            parsed_list[2] = self._run_date.strftime('%Y')
            parsed_list[3] = '1272912'                                    #Vendor number
            parsed_list[4] = price_sheet.cell(row=x, column=10).value     #E-File Code
            parsed_list[5] = ParseFile.parse_item_description(str(price_sheet.cell(row=x, column=1).value))
            parsed_list[6] = 'R'
            parsed_list[8] = 'P' if price_sheet.cell(row=x, column=8).value == '' or price_sheet.cell(row=x, column=8).value == None\
                                    else price_sheet.cell(row=x, column=8).value
            parsed_list[9] = price_sheet.cell(row=x, column=4).value
            parsed_list[10] = price_sheet.cell(row=x, column=2).value
            parsed_list[11] = 'A'
            parsed_list[13] = 'N'
            parsed_list[14] = '20' if price_sheet.cell(row=x, column=5).value == 'Keg' else price_sheet.cell(row=x, column=5).value      # volume
            parsed_list[15] = 'L'
            parsed_list[16] = price_sheet.cell(row=x, column=7).value
            parsed_list[18] = price_sheet.cell(row=x, column=6).value
            parsed_list[19] = ParseFile.get_alcohol_level(0 if price_sheet.cell(row=x, column=9).value == None else
                                                      float(price_sheet.cell(row=x, column=9).value))   #ABV
            parsed_list[21] = str(round(price_sheet.cell(row=x, column=3).value, 2))   #price per bottle use round(x) to round to whole integer value
            parsed_list[22] = str(round(ParseFile.get_case_price(parsed_list[21], parsed_list[16])))
            parsed_list[25] = 'N'
            parsed_list[28] = '1272912'                 #Vendor number
            parsed_list[29] = parsed_list[5]

            #DISCOUNTS
            #2 case
            if price_sheet.cell(row=x, column=24).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '2'
                parsed_list[discount_index+1] = 'C'
                parsed_list[discount_index+2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=24).value))*float(parsed_list[16]),2))
                parsed_list[discount_index+3] = '$'

                discount_index += 4

            #3 case
            if price_sheet.cell(row=x, column=23).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '3'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=23).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'


                discount_index += 4

            #4 case
            if price_sheet.cell(row=x, column=25).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '4'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=25).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 5 case
            if price_sheet.cell(row=x, column=11).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '5'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=11).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 6 case
            if price_sheet.cell(row=x, column=12).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '6'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=12).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4
            # 8 case
            if price_sheet.cell(row=x, column=13).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '8'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=13).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 10 case
            if price_sheet.cell(row=x, column=14).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '10'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=14).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 14 case
            if price_sheet.cell(row=x, column=15).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '14'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=15).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 15 case
            if price_sheet.cell(row=x, column=16).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '15'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=16).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 20 case
            if price_sheet.cell(row=x, column=17).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '20'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=17).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 25 case
            if price_sheet.cell(row=x, column=18).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '25'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=18).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 28 case
            if price_sheet.cell(row=x, column=19).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '28'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=19).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 30 case
            if price_sheet.cell(row=x, column=20).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '30'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=20).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 56 case
            if price_sheet.cell(row=x, column=21).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '56'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=21).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            # 125 case
            if price_sheet.cell(row=x, column=22).value != price_sheet.cell(row=x, column=3).value:
                parsed_list[discount_index] = '125'
                parsed_list[discount_index + 1] = 'C'
                parsed_list[discount_index + 2] = str(round((float(price_sheet.cell(row=x, column=3).value) - float(
                    price_sheet.cell(row=x, column=22).value)) * float(parsed_list[16]), 2))
                parsed_list[discount_index + 3] = '$'

                discount_index += 4

            parsed_array.append(parsed_list)
        else:
            pass

    print 'done parsing'
    ParseFile.write_to_txt(parsed_array)    #write array of parsed rows to output csv file.




  # helper methods
  @staticmethod
  def get_alcohol_level(abv):
      if abv < 7:
          return 'A'
      elif abv >=7 and abv <= 14:
          return 'B'
      elif abv > 14 and abv <= 21:
          return 'C'
      elif abv > 21 and abv <= 24:
          return 'D'

  @staticmethod
  def get_case_price(bottle_price, pack_size):
      return float(bottle_price)*float(pack_size)

  @staticmethod
  def get_discount_bottle_price(bottle_price, pack_size):
      return float(bottle_price) * float(pack_size)


  @staticmethod
  def parse_item_description(item_desc):
      parsed_str = item_desc.rsplit(':', 1)[1]
      return parsed_str


  @classmethod
  def write_to_txt(cls, parsed_data):
      file = open(cls._cwd + '/parseddata/parsed_file_' + str(cls._run_date) + '_wine' +'_retail' '.txt', 'wt')
      for x in parsed_data:
          for y in x:
              file.write('' if y is None else y + '\t')
          file.write('\r')
      file.close()



if __name__ == "__main__":
    parser = ParseFile()
    parser.create_price_file_export()